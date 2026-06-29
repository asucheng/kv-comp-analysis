import os
import openpyxl
from mcp_server.excel_report import TEMPLATE_PATH, load_template
from datetime import date
from mcp_server.models import Subject, Comp, AdjustmentRules, ReportComp, ReportPayload
from mcp_server.estimate import reconcile
from mcp_server.excel_report import load_template, fill_comp_grid, ROWS, apply_ours


def test_template_is_vendored_and_loads():
    assert os.path.isfile(TEMPLATE_PATH)
    wb = load_template()
    assert "Property Comparables" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    # anchor cells the rest of the code depends on
    pc = wb["Property Comparables"]
    assert pc["B6"].value == "Address"
    assert pc["B65"].value == "KV Internal Value"


def _payload():
    s = Subject(address="138 Cranberry Place SE", lat=51.0, lng=-114.0, sqft=1416,
                year_built=2007, beds=3, baths=3, garage=1, community="Cranston",
                property_type="detached")
    comps = [Comp(address=a, lat=51.0, lng=-114.0, sold_price=p, sold_date=date(2026, 4, 1),
                  sqft=sq, year_built=2007, beds=3, baths=3, garage=g, distance_km=d,
                  include_reason="same community")
             for a, p, sq, g, d in [("71 Cranberry Pl", 536_500, 1429, 2, 0.1),
                                    ("78 Cranberry Cl", 560_000, 1425, 2, 0.3),
                                    ("420 Cranberry Cir", 535_000, 1356, 2, 0.2),
                                    ("389 Cranberry Cir", 558_500, 1358, 2, 0.2)]]
    est = reconcile(s, comps, AdjustmentRules(), as_of=date(2026, 6, 10))
    rcomps = [ReportComp(comp=c, kept=True) for c in comps]
    rcomps.append(ReportComp(comp=Comp(address="56 Cranbrook Landing", lat=51.0, lng=-114.0,
                  sold_price=1_173_000, sold_date=date(2026, 4, 1), sqft=1540),
                  kept=False, exclude_reason="lakefront outlier"))
    return ReportPayload(subject=s, comps=rcomps, estimate=est,
                         confidence_reasoning="Tight cluster.", as_of=date(2026, 6, 10))


def _per_comp_by_addr(est, addr):
    return next(ca for ca in est.per_comp if ca.address == addr)


def test_apply_ours_reconciles_grid_to_engine():
    wb = load_template()
    ws = wb["Property Comparables"]
    p = _payload()
    info = fill_comp_grid(ws, p)
    apply_ours(ws, p, info)
    ca = _per_comp_by_addr(p.estimate, "71 Cranberry Pl")   # column E
    # Adjusted Price cell equals the engine's adjusted_price exactly
    assert ws[f"E{ROWS['adjusted_price']}"].value == ca.adjusted_price
    # rows 34..45 sum to Total Adjustments
    total = sum(ws[f"E{r}"].value or 0 for r in range(34, 46))
    assert round(total) == round(ws[f"E{ROWS['total_adj']}"].value)
    # headline: D64 = point/sqft, D65 stamped to point
    assert abs(ws["D64"].value - p.estimate.point / p.subject.sqft) < 1e-6
    assert ws["D65"].value == p.estimate.point
    # range + confidence block written below the headline
    assert ws["C66"].value == p.estimate.low and ws["D66"].value == p.estimate.high
    assert ws["D67"].value == p.estimate.confidence
    # two manual rows relabeled to host our size/time factors
    assert ws[f"B{ROWS['adj_time']}"].value == "Adjustment Time (market)"
    assert ws[f"B{ROWS['adj_size']}"].value == "Adjustment Size (per sqft)"
    # stat range widened past K to the real last comp column (H here)
    assert "E49:H49" in ws["D54"].value


def test_fill_comp_grid_writes_subject_and_all_comps():
    wb = load_template()
    ws = wb["Property Comparables"]
    info = fill_comp_grid(ws, _payload())
    # subject column D
    assert ws[f"D{ROWS['floor_area']}"].value == 1416
    # 4 kept comps land in E,F,G,H
    assert info["cols"] == ["E", "F", "G", "H"]
    assert ws[f"E{ROWS['address']}"].value == "71 Cranberry Pl"
    assert ws[f"E{ROWS['price']}"].value == 536_500
    assert ws[f"E{ROWS['floor_area']}"].value == 1429
    # excluded comp grouped after the kept ones, with a flag
    assert info["excluded_cols"] == ["J"]   # E-H kept, I gap header, J excluded
    assert ws[f"J{ROWS['address']}"].value == "56 Cranbrook Landing"
    # template sample data cleared (K had a sample comp address)
    assert ws[f"K{ROWS['address']}"].value is None
    # formulas captured for Option A reuse
    assert ROWS["adjusted_price"] in info["formulas"]
