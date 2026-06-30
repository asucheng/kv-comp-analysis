from __future__ import annotations
import io
import os
import warnings
import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formula.translate import Translator
from mcp_server.models import ReportPayload, Comp, Subject

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "templates", "sf_uw_template.xlsx")


def load_template() -> "openpyxl.Workbook":
    """Load a fresh, writable copy of the KV underwriter template. Read-only on disk —
    callers mutate the returned in-memory workbook and save elsewhere."""
    if not os.path.isfile(TEMPLATE_PATH):
        raise FileNotFoundError(f"KV Excel template missing at {TEMPLATE_PATH}")
    with warnings.catch_warnings():
        # openpyxl warns that the template's Data Validation extension is dropped; harmless.
        warnings.simplefilter("ignore")
        return openpyxl.load_workbook(TEMPLATE_PATH)


ROWS = {
    "address": 6, "source": 7, "price": 9, "list_sale": 11, "neighbourhood": 13,
    "unit_price": 14, "sale_date": 15, "floor_area": 17, "site_size": 18, "style": 19,
    "format": 20, "year_built": 21, "beds": 22, "baths": 23, "basement": 24,
    "parking": 25, "fireplace": 26, "deck": 27, "appliances": 28, "notes": 29,
    "appraised_list": 30, "bidask": 31, "adj_time": 34, "adj_size": 35, "adj_garage": 36,
    "adj_beds": 37, "adj_baths": 38, "adj_finishing": 39, "adj_basement": 40,
    "adj_appliances": 41, "adj_fireplace": 42, "adj_neighbourhood": 43, "adj_deck": 44,
    "adj_other": 45, "total_adj": 46, "adjusted_price": 48, "adjusted_unit": 49,
}
# Rows that hold a per-comp formula in column E of the template (reused for Option A).
FORMULA_ROWS = (ROWS["unit_price"], ROWS["appraised_list"], ROWS["bidask"],
                ROWS["adj_time"], ROWS["adj_size"], ROWS["adj_garage"], ROWS["adj_beds"],
                ROWS["adj_baths"], ROWS["adj_finishing"], ROWS["adj_basement"],
                ROWS["adj_appliances"], ROWS["adj_fireplace"], ROWS["adj_neighbourhood"],
                ROWS["adj_deck"], ROWS["adj_other"], ROWS["total_adj"],
                ROWS["adjusted_price"], ROWS["adjusted_unit"])

_FIRST_COMP_COL = "E"          # subject is D; comps start at E
_CLEAR_LAST_COL = "P"          # sample comps + stray notes live in E..P
_CLEAR_ROWS = range(5, 50)     # attribute + adjustment block
# KV's template has 7 comp slots; live data keeps 100+ comps, so the grid would sprawl to
# 100+ columns. Display only the closest N as columns (all comps still drive the engine's
# value, which is written independently into the headline). A note discloses the full count.
_MAX_COMP_COLS = 7
_COMP_COUNT_NOTE_ROW = 50      # blank row just below Adjusted Unit Price (49), above stats (51)


def _set(ws, col: str, row: int, val) -> None:
    ws[f"{col}{row}"] = val


def _capture_formulas(ws) -> dict:
    """Read column-E formula strings before clearing, so Option A can re-instantiate them
    for every comp column (including ones past the template's K)."""
    out = {}
    for r in FORMULA_ROWS:
        v = ws[f"E{r}"].value
        if isinstance(v, str) and v.startswith("="):
            out[r] = v
    return out


def _clear_comp_region(ws) -> None:
    first = column_index_from_string(_FIRST_COMP_COL)
    last = column_index_from_string(_CLEAR_LAST_COL)
    for r in _CLEAR_ROWS:
        for ci in range(first, last + 1):
            ws.cell(row=r, column=ci).value = None


def _fill_attr_col(ws, col: str, c: Comp, *, source: str = "HD") -> None:
    _set(ws, col, ROWS["address"], c.address)
    _set(ws, col, ROWS["source"], source)
    _set(ws, col, ROWS["price"], c.sold_price)
    _set(ws, col, ROWS["list_sale"], "Sale Price")
    sd = ws[f"{col}{ROWS['sale_date']}"]
    sd.value = datetime(c.sold_date.year, c.sold_date.month, c.sold_date.day)
    sd.number_format = "m/d/yyyy"
    _set(ws, col, ROWS["floor_area"], c.sqft)
    if c.year_built is not None:
        _set(ws, col, ROWS["year_built"], c.year_built)
    if c.beds is not None:
        _set(ws, col, ROWS["beds"], c.beds)
    if c.baths is not None:
        _set(ws, col, ROWS["baths"], c.baths)
    if c.parking_type:
        _set(ws, col, ROWS["parking"], c.parking_type)
    if c.style:
        _set(ws, col, ROWS["style"], c.style)
    if c.basement:
        _set(ws, col, ROWS["basement"], c.basement)
    if c.community:
        _set(ws, col, ROWS["neighbourhood"], c.community)
    if c.include_reason:
        _set(ws, col, ROWS["notes"], c.include_reason)


def _fill_subject_col(ws, s: Subject) -> None:
    _set(ws, "D", ROWS["address"], s.resolved_address or s.address)
    if s.community:
        _set(ws, "D", ROWS["neighbourhood"], s.community)
    if s.sqft is not None:
        _set(ws, "D", ROWS["floor_area"], s.sqft)
    if s.lot_sf is not None:
        _set(ws, "D", ROWS["site_size"], s.lot_sf)
    if s.year_built is not None:
        _set(ws, "D", ROWS["year_built"], s.year_built)
    if s.beds is not None:
        _set(ws, "D", ROWS["beds"], s.beds)
    if s.baths is not None:
        _set(ws, "D", ROWS["baths"], s.baths)
    if s.parking_type:
        _set(ws, "D", ROWS["parking"], s.parking_type)


def fill_comp_grid(ws, payload: ReportPayload, max_comp_cols: int = _MAX_COMP_COLS) -> dict:
    kept = [rc for rc in payload.comps if rc.kept]
    kept.sort(key=lambda rc: (rc.comp.distance_km is None, rc.comp.distance_km or 0))
    total_kept = len(kept)
    shown = kept[:max_comp_cols]            # display the closest N; all kept comps stay in the math
    excluded = [rc for rc in payload.comps if not rc.kept]

    formulas = _capture_formulas(ws)
    _clear_comp_region(ws)
    _fill_subject_col(ws, payload.subject)

    cols: list[str] = []
    idx = column_index_from_string(_FIRST_COMP_COL)
    for rc in shown:
        col = get_column_letter(idx)
        _fill_attr_col(ws, col, rc.comp)
        cols.append(col)
        idx += 1

    excluded_cols: list[str] = []
    if excluded:
        idx += 1  # one-column gap separates kept from excluded
        header_col = get_column_letter(idx - 1)
        ws[f"{header_col}4"] = "Excluded"
        for rc in excluded:
            col = get_column_letter(idx)
            _fill_attr_col(ws, col, rc.comp)
            if rc.exclude_reason:
                _set(ws, col, ROWS["notes"], rc.exclude_reason)
            excluded_cols.append(col)
            idx += 1

    if total_kept > len(shown):
        ws[f"B{_COMP_COUNT_NOTE_ROW}"] = (
            f"Showing {len(shown)} of {total_kept} comps (closest by distance); "
            f"all {total_kept} were used in the valuation.")

    last_col = get_column_letter(idx - 1)
    return {"cols": cols, "excluded_cols": excluded_cols,
            "last_col": last_col, "formulas": formulas,
            "kept_addresses": [rc.comp.address for rc in shown],
            "total_kept": total_kept, "shown": len(shown)}


# ---------------------------------------------------------------------------
# Our-math adjustments + headline value (method="ours")
# ---------------------------------------------------------------------------

# factor -> the template row it occupies under our-math.
_FACTOR_ROW = {
    "time": ROWS["adj_time"], "size": ROWS["adj_size"], "garage": ROWS["adj_garage"],
    "beds": ROWS["adj_beds"], "full_baths": ROWS["adj_baths"], "half_baths": ROWS["adj_baths"],
    # Our method derives a year-built adjustment KV's template has no dedicated row for; host it in
    # the "Other" row (relabeled in apply_ours) so itemized rows still reconcile with Total Adjustments.
    "year_built": ROWS["adj_other"],
}

def _adj_dollars(ca) -> dict:
    """Per-factor dollar impact for one comp, summing to adjusted_price - raw_price.
    Time is a pct on the raw price; everything else is already a dollar amount."""
    out: dict[int, float] = {}
    for a in ca.adjustments:
        row = _FACTOR_ROW.get(a.factor)
        if row is None:
            continue
        d = ca.raw_price * a.value_pct if a.value_pct is not None else (a.value_dollar or 0.0)
        out[row] = out.get(row, 0.0) + d   # full+half baths fold into the Bathroom row
    return out


def apply_ours(ws, payload: ReportPayload, info: dict) -> None:
    ws[f"B{ROWS['adj_time']}"] = "Adjustment Time (market)"
    ws[f"B{ROWS['adj_size']}"] = "Adjustment Size (per sqft)"
    ws[f"B{ROWS['adj_other']}"] = "Adjustment Year Built"
    by_addr = {ca.address: ca for ca in payload.estimate.per_comp}

    # Map columns to comps using the addresses fill_comp_grid recorded (in grid order),
    # so column E here is always the same comp as column E in the grid — no separate sort.
    for col, addr in zip(info["cols"], info["kept_addresses"]):
        ca = by_addr.get(addr)
        if ca is None:
            continue
        _set(ws, col, ROWS["appraised_list"], ca.raw_price)
        _set(ws, col, ROWS["bidask"], "$nil")
        dollars = _adj_dollars(ca)
        for row in range(34, 46):
            _set(ws, col, row, round(dollars.get(row, 0.0)))
        total = round(ca.adjusted_price - ca.raw_price)
        _set(ws, col, ROWS["total_adj"], total)
        _set(ws, col, ROWS["adjusted_price"], ca.adjusted_price)
        _set(ws, col, ROWS["adjusted_unit"], ca.adjusted_ppsf)

    _write_headline(ws, payload)
    _widen_stat_ranges(ws, info["cols"][-1] if info["cols"] else "K")


def _write_headline(ws, payload: ReportPayload) -> None:
    est, s = payload.estimate, payload.subject
    if s.sqft:
        ws["D64"] = est.point / s.sqft       # D65 = D64*D17 = point (on recalc)
    ws["D65"] = est.point                     # stamp static so non-Excel viewers are correct
    ws["B66"] = "KV Value Range (25th-75th)"
    # One string in D66 (not low in the narrow C "units" column, which renders as "###").
    ws["D66"] = f"{est.low:,.0f} - {est.high:,.0f}"
    ws["B67"] = "Confidence"
    ws["D67"] = est.confidence


def _widen_stat_ranges(ws, last_col: str) -> None:
    """Repoint the unit-price stat formulas from the template's fixed E:K to E:<last_col>."""
    rng = f"E49:{last_col}49"
    ws["D54"] = f"=MIN({rng})"
    ws["D55"] = f"=_xlfn.PERCENTILE.INC({rng},0.25)"
    ws["D56"] = f"=AVERAGE({rng})"   # D56 is labelled "Median" in KV's template but uses AVERAGE — faithfully reproduced, not a bug
    ws["D58"] = f"=_xlfn.PERCENTILE.INC({rng},0.75)"
    ws["D59"] = f"=MAX({rng})"
    ws["D60"] = f"=_xlfn.STDEV.P({rng})"
    ws["D57"] = f"=SUM(E48:{last_col}48)/SUM(E17:{last_col}17)"


# ---------------------------------------------------------------------------
# Template-math path (method="template") — Option A
# ---------------------------------------------------------------------------


def apply_template(ws, payload: ReportPayload, info: dict) -> None:
    """Option A: feed our derived coefficients into KV's Table A and let the sheet's own
    formulas compute. Re-instantiate the per-comp formulas (captured before clearing) for
    every comp column so they exist past the template's original K."""
    coeffs = {c.factor: c.value for c in payload.estimate.coefficients}
    if "beds" in coeffs:
        ws["D73"] = coeffs["beds"]
    if "full_baths" in coeffs:
        ws["D74"] = coeffs["full_baths"]   # only the full-bath coeff feeds D74; the template's bath formula uses a combined count, so the half-bath delta is intentionally approximated away in this opt-in path
    if "garage" in coeffs:
        ws["D75"] = coeffs["garage"]

    formulas = info["formulas"]
    for col in info["cols"] + info["excluded_cols"]:
        for row, f in formulas.items():
            ws[f"{col}{row}"] = Translator(f, origin=f"E{row}").translate_formula(f"{col}{row}")
    _widen_stat_ranges(ws, info["cols"][-1] if info["cols"] else "K")


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

_SUMMARY_BLANK = ("D10", "D11", "D12", "D22", "D24", "D37", "D42")


def fill_summary(wb, payload: ReportPayload) -> None:
    s = payload.subject
    sm = wb["Summary"]
    sm["B3"] = s.resolved_address or s.address
    if s.community:
        sm["D14"] = s.community
    if s.year_built is not None:
        sm["D16"] = s.year_built
    if s.sqft is not None:
        sm["D20"] = s.sqft
    if s.lot_sf is not None:
        sm["D18"] = round(s.lot_sf / 10.7639, 2)   # m2; D19 = D18*10.7639 recomputes sqft
    for cell in _SUMMARY_BLANK:
        sm[cell] = None
    if s.hd_estimate:
        sm["B46"] = "HD AVM (cross-check, not a sale)"
        sm["D46"] = s.hd_estimate


def render_report_xlsx(payload: ReportPayload, method: str = "ours") -> bytes:
    if method not in ("ours", "template"):
        raise ValueError(f"unknown method {method!r}; expected 'ours' or 'template'")
    wb = load_template()
    ws = wb["Property Comparables"]
    info = fill_comp_grid(ws, payload)
    if method == "ours":
        apply_ours(ws, payload, info)
    else:
        apply_template(ws, payload, info)
    fill_summary(wb, payload)
    try:
        wb.calculation.fullCalcOnLoad = True   # force Excel to recompute remaining formulas
    except Exception:
        pass
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
